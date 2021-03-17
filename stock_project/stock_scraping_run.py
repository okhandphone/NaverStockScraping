from scraping.stock_scraping_master import get_market_fluctuation, get_theme_fluctuation
# mk_code_dict = get_naver_market_code()
# get_share_code_from_naver(mk_code_dict)
# get_realtime_value(mk_code_dict)

# code_dict = get_code_from_excel()
# code_dict = {'352940': {'Share Name': '인바이오 *', 'Market Code': '140', 'Market Name': '화학'}, '357550': {'Share Name': '석경에이티 *', 'Market Code': '140', 'Market Name': '화학'}, '020120': {'Share Name': '키다리스튜디오 ', 'Market Code': '236', 'Market Name': '양방향미디어와서비스'}, '239340': {'Share Name': '줌인터넷 *', 'Market Code': '236', 'Market Name': '양방향미디어와서비스'}, '300080': {'Share Name': '플리토 *', 'Market Code': '236', 'Market Name': '양방향미디어와서비스'}, '035420': {'Share Name': 'NAVER ', 'Market Code': '236', 'Market Name': '양방향미디어와서비스'}}
# # get_financial_info(code_dict)
# get_market_fluctuation()
# get_theme_fluctuation()

# mk_code_dict = get_naver_market_code()
# get_real_time_value(mk_code_dict)

import os
import csv
import time
import datetime
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from scraping.stock_scraping_master import get_naver_market_code
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
import pandas as pd

'''
# 1. CSV 파일 생성
'''
dir_path = "stock_data/실시간벨류에이션"
# if not os.path.exists(dir_path):
#     os.mkdir(dir_path)
# csv_file = open(dir_path + f'\\realtime_stock_value_{datetime.date.today()}.csv', 'w', encoding='utf-8-sig', newline='')
# csv_writer = csv.writer(csv_file)
# csv_writer.writerow(['URL', '종목코드', '종목명', '업종코드', '업종명', '등락률', '시가총액 (억 원)', 'PER', 'ROE', 'PEG', 'ROA', 'PBR', '유보율'])
#
# '''
# # 2. 네이버 업종 크롤링
# '''
# # headless
# options = webdriver.ChromeOptions()
# options.headless = True
# options.add_argument("window-size=1920x1080")
# options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
# driver = webdriver.Chrome(options=options)
#
# # 마켓코드 받아오기
# mk_code_dict = get_naver_market_code()
#
# for mk_name, mk_code in mk_code_dict.items():
#     # try:
#     # time.sleep(3)
#     print(f"Start {mk_name} Category")
#     url = f'https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no={mk_code}'
#
#     # 셀레니움으로 받아야 옵션 정보가 유지됨
#     driver.get(url)
#     driver.implicitly_wait(3)
#     wait = WebDriverWait(driver, 3)
#
#     titles = driver.find_elements_by_css_selector("#contentarea > div:nth-child(5) > table > thead > tr:nth-child(1) > th")
#     title = [line.text for line in titles]
#     print(title, "sel")
#
#     # title에 '거래량'이 있을 경우 옵션변경
#     if '거래량' in title:
#
#         # 기존 옵션 제거 : 거래량, 매수호가, 거래대금, 매도호가, 전일거래량
#         remove_list = [1, 2, 3, 8, 9]
#         for num in remove_list:
#             wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#option{num}'))).click()
#
#         # 원하는 옵션 클릭 : 시가총액, PER, ROE, ROA, PBR, 유보율
#         remove_list = [4, 6, 12, 18, 24, 27]
#         for num in remove_list:
#             wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#option{num}'))).click()
#
#         # 옵션 적용 클릭
#         wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.item_btn > a'))).click()  # 적용하기
#         time.sleep(1)
#
#     # html 파싱
#     soup = BeautifulSoup(driver.page_source, "lxml")
#     # 타이틀 다시 확인
#     headers = soup.select('#contentarea > div:nth-child(5) > table > thead > tr:nth-child(1) > th')
#     head = [line.text for line in headers]
#     print(title, "soup")
#     table = soup.select('#contentarea > div.box_type_l')[1].select('tbody > tr')[:-2]
#
#     # 종목별 정보 추출 및 기록
#     for line in table:
#         # 리스트에 이름 및 코드 추가
#         share_name = line.td.text
#         share_code = line.td.a['href'].replace("/item/main.nhn?code=", "")
#         share_link = "https://finance.naver.com/item/main.nhn?code=" + share_code
#         info_list = [share_link, str(share_code), share_name, str(mk_code), mk_name]
#
#         # 벨류 데이터 추가 (등락률 부터)
#         data = line.select('td')[3:-1]
#         for num in data:
#             num = num.text.replace(',', '').replace('+', '').replace('%', '')
#             if num == '':
#                 num = ''
#             else:
#                 num = float(num)
#             info_list.append(num)
#
#         # PEG 밸류 추가
#         per = info_list[7]
#         roe = info_list[8]
#         # per, roe 값을 기준으로 info_list 값 달라짐
#         if per == '' or roe == '':  # 값이 없는 경우
#             info_list.insert(9, '')
#         elif per > 0 and roe > 0:  # 0보다 크거나 같으면 peg 계산
#             peg = per / roe
#             info_list.insert(9, f"{peg:.1f}")
#         elif per <= 0 or roe <= 0:
#             info_list.insert(9, '')  # 마이너스인 경우 '' 반환
#
#         # csv에 기록
#         csv_writer.writerow(info_list)
#
#     # except:
#     #     print(f"err code: {mk_code}, {mk_name}")
#     #     pass
#
# driver.quit()
# csv_file.close()
# print("CSV Completed")


'''
# 3. realtime_stock_value.xlsx 파일에 날짜 이름 시트로 붙여넣기
'''
file_name = dir_path + "/realtime_stock_value.xlsx"
# df = pd.read_csv(dir_path + f'/realtime_stock_value_{datetime.date.today()}.csv')
#
# if not os.path.exists(file_name):
#     with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name=f"{datetime.date.today()}",index=False)
# else:
#     with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name=f"{datetime.date.today()}",index=False)
# print("df to 엑셀")


'''
# 4. 엑셀 서식 추가
'''
# 셀서식
align_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 글꼴 서식
font_bold = Font(name="맑은 고딕", size=11, bold=True)
font_red = Font(name="맑은 고딕", size=11, color="FE2E2E")
font_blue = Font(name="맑은 고딕", size=11, color="1E88E5")

# color_fill
orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("EC7063"))
dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("5DADE2"))


# 워크북 불러오기
wb = load_workbook(file_name)
ws = wb[f"{datetime.date.today()}"]

# 컬럼너비 (종목명, 업종명, 상승종목 수, 하락종목 수)
width_dict = {"C": 21, "E": 19, "O": 13, "P": 13, "Q": 13}
for key, value in width_dict.items():
    ws.column_dimensions[f"{key}"].width = value

# 헤더 서식
for row in ws["A1:M1"]:
    for cell in row:
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = orange_fill
        cell.border = thin_border

# 등락률 글자색
count_inc = 0 # 상승 종목 갯수
count_dec = 0 # 하락 종목 갯수
for col in ws[f"F2:F{ws.max_row}"]:
    for cell in col:
        if cell.value > 0:
            cell.font = font_red
            count_inc += 1
        else:
            cell.font = font_blue
            count_dec += 1

# 하이퍼링크
for i in range(2, ws.max_row):
    ws[f"C{i}"].hyperlink = ws[f"A{i}"].value

# 상승 하락 종목 셀 서식
ws["O1"].value = "상승 종목 수"
ws["P1"].value = "하락 종목 수"
ws["Q1"].value = "총 종목 수"
ws["O1"].fill = dark_pink_fill
ws["P1"].fill = dark_blue_fill
ws["Q1"].fill = orange_fill
ws["O2"].value = count_inc
ws["P2"].value = count_dec
ws["Q2"].value = count_inc + count_dec

for row in ws[f"O1:Q1"]:
    for cell in row:
        cell.font = font_bold

for row in ws[f"O1:Q2"]:
    for cell in row:
        cell.border = thin_border
        cell.alignment = align_center

print(count_inc, count_dec, count_inc + count_dec)

# 전체 보더
for row in ws[f"A2:M{ws.max_row}"]:
    for cell in row:
        cell.border = thin_border

# 오토필터
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'

# 컬럼 숨기기 (URL)
ws.column_dimensions["A"].hidden= True

wb.save(file_name)
print("엑셀 서식 완료")
