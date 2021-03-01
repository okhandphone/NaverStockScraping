from openpyxl import load_workbook  #  파일 불러오기

wb = load_workbook("sample03.xlsx")  #  불러올 파일명 넣기
ws = wb.active  #  활성화된 시트 가져오기

# 테마 코드 리스트
th_code_list = []
for x in range(2, ws.max_row + 1):
    th_code = int(ws.cell(row=x, column=1).value)
    th_code_list.append(th_code)
print(th_code_list)
