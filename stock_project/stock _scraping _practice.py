# This Python file uses the following encoding: utf-8
import os, sys

import requests
from bs4 import BeautifulSoup
import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


# # 뉴스 저장 경로 및 파일 만들기
# news_path = "stock_data\\News"
# if not os.path.exists(news_path + "\\데일리 뉴스.xlsx"):
#     os.mkdir(news_path)
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "데일리 뉴스"
#     ws.append(['링크', '키워드', '날짜', '타이틀', '요약'])
#
#     # 속성 {열번호: 너비}
#     attrs = {"B": "12", "C": "11", "D": "49"}
#     for key, value in attrs.items():
#         ws.column_dimensions[f"{key}"].width = f"{value}"
#
#     # for i in range(1, ws.max_column):
#     #     ws[f"A{i}"].font.copy(bold=True)
#     # ws.rows(1).alignment = Alignment(ho)
#     wb.save(news_path + "\\데일리 뉴스.xlsx")
#     print("새 엑셀 파일 생성")
#
# # 이미 디렉토리가 있다면 엑셀 파일 로드
# else:
#     wb = load_workbook(news_path + "\\데일리 뉴스.xlsx")
#     ws = wb.active
#     print("데일리뉴스 엑셀파일 로드")

# 검색학 뉴스키워드
while True:
    keyword = input("키워드 입력 : ")
    # keyword = ['sk', '현대차', '현대모비스', '기아차', '미국 증시', '연준', 'FOMC', '끝']
    # 끝이라고 입력하면 프로그램 종료
    if keyword == "끝":
        break

    # # 이미지 폴더 만들기
    # img_path = f"stock_data\\News\\image"
    # if not os.path.exists(img_path):
    #     os.mkdir(img_path)

    # 오늘날짜 함께 입력
    today = datetime.date.today()
    print(today)
    # 뉴스 페이지 '최신순' 첫번재 페이지
    url = f"https://search.naver.com/search.naver?where=news&query={keyword}&sm=tab_srt&sort=1&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so%3Add%2Cp%3Aall%2Ca%3Aall&mynews=0&refresh_start=0&related=0"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.72 Safari/537.36 Edg/89.0.774.45"}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    t = soup.select('.news_area > a')
    img = soup.select('.news_wrap .dsc_thumb img')

    for i in range(len(t)):

        # 타이틀 및 링크
        title = t[i].text
        link = t[i]['href']
        # 내용 받아오기
        news_req = requests.get(link).text
        news_soup = BeautifulSoup(news_req, "html.parser")
        content = news_soup.select_one("subTitle_s2 br").text.strip()
        print(f"제목 : {title}")
        print(f"링크 : {link}")
        print(content)
        print()
        # ws.append([link, keyword, today, title])
        # wb.save(news_path + "\\데일리 뉴스.xlsx")

        # # 이미지도 함께 다운 / 입력
        # img_url = img[i]['src']
        # response = requests.get(img_url)
        # response.raise_for_status()
        #
        # with open(img_path + f"\\{keyword}_{i + 1}.png", "wb") as f:
        #     f.write(response.content)
        #
        # img_for_excel = Image(img_path + f"\\{keyword}_{i + 1}.png")
        # insert_img = ws.add_image(img_for_excel, f"A{ws.max_row}")
        # ws.append([insert_img, link, keyword, today, title])


# for i in range(len(t)):
    #     try:
    #         # 타이틀 및 링크
    #         title = t[i].text
    #         link = t[i]['href']
    #         # 내용 받아오기
    #         news_res = requests.get(link)
    #         news_soup = BeautifulSoup(news_res, "html.parser")
    #         content = news_soup.select_one("subTitle_s2 br").text.strip()
    #         print(f"제목 : {title}")
    #         print(f"링크 : {link}")
    #         print(content)
    #         print()
    #         # ws.append([link, keyword, today, title])
    #         # wb.save(news_path + "\\데일리 뉴스.xlsx")
    #
    #         # # 이미지도 함께 다운 / 입력
    #         # img_url = img[i]['src']
    #         # response = requests.get(img_url)
    #         # response.raise_for_status()
    #         #
    #         # with open(img_path + f"\\{keyword}_{i + 1}.png", "wb") as f:
    #         #     f.write(response.content)
    #         #
    #         # img_for_excel = Image(img_path + f"\\{keyword}_{i + 1}.png")
    #         # insert_img = ws.add_image(img_for_excel, f"A{ws.max_row}")
    #         # ws.append([insert_img, link, keyword, today, title])
    #     except:
    #         pass



# mk_code_dict = {'증권': '12', '건설': '42', '손해보험': '190', '디스플레이장비및부품': '199', '섬유,의류,신발,호화품': '134'}
# get_real_time_value(mk_code_dict)

# data = [77,
# 14745,
# 6380,
# 6390,
# 3486914,
# 22755,
# 5129739,
# 20200,
# 20250,
# 7343076,
# 155598,
# 6702288]
# #
# # for num in data:
# #     print(num)
#
# [print(i) for i in data]

# get_share_code()의 리턴값 코드 딕셔너리를 받아서 연별, 분기별 실적 정리
# code_dict = {'016610': {'Share Name': 'DB금융투자 ', 'Market Code': '12', 'Market Name': '증권'}, '375500': {'Share Name': 'DL이앤씨 ', 'Market Code': '42', 'Market Name': '건설'}, '112190': {'Share Name': 'DB손해보험 ', 'Market Code': '190', 'Market Name': '손해보험'},'068790': {'Share Name': 'DMS *', 'Market Code': '199', 'Market Name': '디스플레이장비및부품'}, '001530': {'Share Name': 'DI동일 ', 'Market Code': '134', 'Market Name': '섬유,의류,신발,호화품'}}
# import time
# from bs4 import BeautifulSoup
# from pandas import DataFrame
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
#
# # headless webdriver
# options = webdriver.ChromeOptions()
# options.headless = True
# options.add_argument("window-size=1920x1080")
# options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
# driver = webdriver.Chrome(options=options)
# driver.maximize_window()
# driver.implicitly_wait(3)
#
# # header = ",Share Code,Share Name,Market Code,Market Name,매출액,영업이익,영업이익(발표기준),세전계속사업이익,당기순이익,  당기순이익(지배),  당기순이익(비지배),자산총계,부채총계,자본총계,  자본총계(지배),  자본총계(비지배),자본금,영업활동현금흐름,투자활동현금흐름,재무활동현금흐름,CAPEX,FCF,이자발생부채,영업이익률,순이익률,ROE(%),ROA(%),부채비율,자본유보율,EPS(원),PER(배),BPS(원),PBR(배),현금DPS(원),현금배당수익률,현금배당성향(%),발행주식수(보통주)".split(',')
#
# print("Get Year Financial Info Start : ")
# print(time.strftime("%Y-%m-%d %H:%M:%S"))
# code_dict = {'016610': {'Share Name': 'DB금융투자 ', 'Market Code': '12', 'Market Name': '증권'}, '005830': {'Share Name': 'DB손해보험 ', 'Market Code': '190', 'Market Name': '손해보험'}, '112190': {'Share Name': 'KC산업 ', 'Market Code': '191', 'Market Name': '상사'}, '000990': {'Share Name': 'DB하이텍 ', 'Market Code': '202', 'Market Name': '반도체와반도체장비'}, '000995': {'Share Name': 'DB하이텍1우 ', 'Market Code': '202', 'Market Name': '반도체와반도체장비'}, '139130': {'Share Name': 'DGB금융지주 ', 'Market Code': '20', 'Market Name': '은행'}, '001530': {'Share Name': 'DI동일 ', 'Market Code': '134', 'Market Name': '섬유,의류,신발,호화품'}, '000210': {'Share Name': 'DL ', 'Market Code': '42', 'Market Name': '건설'}, '000215': {'Share Name': 'DL우 ', 'Market Code': '42', 'Market Name': '건설'}, '375500': {'Share Name': 'DL이앤씨 ', 'Market Code': '42', 'Market Name': '건설'}, '068790': {'Share Name': 'DMS *', 'Market Code': '199', 'Market Name': '디스플레이장비및부품'}, '112190': {'Share Name': 'DRB동일 ', 'Market Code': '36', 'Market Name': '자동차부품'}}
#
# for key, value in code_dict.items():
#     try:
#         time.sleep(3)
#         print(f"Start: {key}")
#         driver.get(f'https://finance.naver.com/item/coinfo.nhn?code={key}')
#         wait = WebDriverWait(driver, 3)
#
#         # iframe 접근
#         driver.switch_to.frame('coinfo_cp')
#         time.sleep(1.5)
#
#         # 연간 분기 돌아가면서 정보 받아오기
#         period_dict = {'cns_Tab21': '연간', 'cns_Tab22': '분기'}
#         for p_key, p_val in period_dict.items():
#             wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#{p_key}'))).click()
#             time.sleep(1)
#
#             # iframe 에서 html정보 추출
#             soup = BeautifulSoup(driver.page_source, "html.parser")
#
#             # 기간 정보
#             period_data = soup.select('table.gHead01')[3].select('thead > tr > th')[2:7]
#             period = [line.text.strip()[:7] for line in period_data]
#
#             # financial_dict에 코드정보 입력
#             financial_dict = {'Share Code': f'{str(key)}'}
#             financial_dict.update(value)  # code_dict의 정보를 추가
#
#             # 재무정보
#             financial_table = soup.select('table.gHead01')[3].select('tbody > tr')
#             for line in financial_table:
#                 name = line.select_one('th').text  # 재무 항목명
#                 data = line.select('td')[:5]  # 재무 데이터 : 앞 다섯 기간만, 컨센 제외
#                 financial_dict.setdefault(name, [num.text.replace(",", "") for num in data])
#
#             # financial_dict를 데이터 프레임으로
#             # 컬럼 = list(financial_dict.keys()) = name, 인덱스 = 기간정보
#             financial_df = DataFrame(financial_dict, columns=list(financial_dict.keys()), index=period)
#             financial_df.to_csv(f"{p_val}실적_raw.csv", mode="a", header=False, encoding='utf-8-sig')
#     except:
#         pass
#         print(f'pass: {key}')
# print("Financial Info Done")
# print(time.strftime("%Y-%m-%d %H:%M:%S"))
# driver.quit()


'''def get_quarter_financial_info(code_dict):

    from bs4 import BeautifulSoup
    from pandas import DataFrame
    from selenium import webdriver
    import time

    # headless
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument("window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.implicitly_wait(3)

    # get_share_code()의 리턴값 코드 딕셔너리를 받아서 분기 정보를 정리

    quarter_financial_dict = {}
    for key in code_dict.keys():
        driver.get(f'https://finance.naver.com/item/coinfo.nhn?code={key}')

        # iframe 에서 html정보 추출
        driver.switch_to.frame('coinfo_cp')

        # 분기 정보
        time.sleep(1.0)
        driver.find_element_by_id('cns_Tab22').click()
        time.sleep(0.5)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # 기간 정보
        period_data = soup.select('table.gHead01')[3].select('thead > tr > th')[2:7]
        period_quarter = []
        for line in period_data:
            date = line.text.strip()[:7]
            period_quarter.append(date)
        # print(period_quarter)

        # 재무정보
        financial_table = soup.select('table.gHead01')[3].select('tbody > tr')
        # code_dict의 정보를 추가하여 financial _dict 생성
        # quarter_financial_dict = {"Share Code": f'{key}',
        #                           'Share Name': code_dict[f'{key}']['Share Name'],
        #                           'Market Code': code_dict[f'{key}']['Market Code'],
        #                           'Market Name': code_dict[f'{key}']['Market Name']}
        for line in financial_table:
            name = line.select_one('th').text
            data = line.select('td')[:5]
            # 숫자는 실수로 변경
            for num in data:
                num = num.text.replace(",", "")
                if num == 'N/A':
                    num = ''
                elif num == '':
                    num = ''
                else:
                    num = float(num)
                # financial_dict : key = name, value = data num list
                quarter_financial_dict.setdefault(name, []).append(num)

        # financial_dict를 데이터 프레임으로 : 컬럼 = financial_dict.keys() = name, 로우 인덱스 = 기간정보
        quarter_financial_df = DataFrame(quarter_financial_dict, columns=list(quarter_financial_dict.keys()), index=period_quarter)
        # quarter_financial_df.index.name = 'Period'

    driver.quit()
    # return quarter_financial_dict

# print(get_quarter_financial_info(code_dict))



# 3. 연별 실적 스크래핑

def get_year_financial_info(code_dict):
    from bs4 import BeautifulSoup
    from pandas import DataFrame
    from selenium import webdriver
    import time

    # headless
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument("window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    driver.implicitly_wait(3)


    print("Get Year Financial Info Start")
    # get_share_code()의 리턴값 코드 딕셔너리를 받아서 연별실적 정리
    for key, value in code_dict.items():
        driver.get(f'https://finance.naver.com/item/coinfo.nhn?code={key}')

        # iframe 에서 html정보 추출
        driver.switch_to.frame('coinfo_cp')

        # 연간 정보 접근
        time.sleep(1)
        driver.find_element_by_id('cns_Tab21').click()
        time.sleep(0.5)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # 기간 정보
        period_data = soup.select('table.gHead01')[3].select('thead > tr > th')[2:7]
        period_year = []
        for line in period_data:
            date = line.text.strip()[:4]
            period_year.append(date)

        # 재무정보
        financial_table = soup.select('table.gHead01')[3].select('tbody > tr')

        # financial _dict에 기간별 재무정보 입력
        year_financial_dict = {'Share Code': f'{str(key)}'}
        year_financial_dict.update(value)  # code_dict의 정보를 추가

        for line in financial_table:
            name = line.select_one('th').text  # 재무 항목명
            data = line.select('td')[:5]  # 재무 데이터
            for num in data:
                num = num.text.replace(",", "")
                # financial_dict 구성 : key = name, value = num list
                year_financial_dict.setdefault(name, []).append(num)

        # financial_dict를 데이터 프레임으로 : 컬럼 = financial_dict.keys() = name, 로우 인덱스 = 기간정보
        year_financial_df = DataFrame(year_financial_dict, index=period_year)
        year_financial_df.to_csv("연별실적.csv", mode="a", header=False, encoding='utf-8-sig')
        # year_financial_df.index.name = 'Period'
        # print(year_financial_df)
    driver.quit()
    print("Fianancial Info Done")

    # return year_financial_dict


# from kyu_def.stock_scraping_master import get_etf_info
# etf_code = ['069500', '367380', '341850', '376250', '245350']

import time
import datetime
import csv
from kyu_def.web_scraping import create_soup
from selenium import webdriver
from bs4 import BeautifulSoup

# # CSV
# csv_file = open(f'etf_{datetime.date.today()}.csv', 'a', encoding='utf-8-sig', newline='')
csv_writer = csv.writer(csv_file)
csv_writer.writerow(['ETF이름', '링크', 'ETF코드', '운용사', '수수료', '시가총액(억 원)', '수익률', '', '', '', '구성종목 TOP 10'])
csv_writer.writerow(['', '', '', '', '', '', '1개월', '3개월', '6개월', '1년', '1위', '', '2위', '', '3위', '', '4위', '', '5위', '', '6위', '','7위', '', '8위', '', '9위', '', '10위', ''])

# 헤드리스 셀레니움
options = webdriver.ChromeOptions()
options.headless = True
options.add_argument("window-size=1920x1080")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(3)

# print(f"Get ETF info :")
# print(time.strftime("%Y-%m-%d %H:%M:%S"))


url = 'https://finance.naver.com/item/coinfo.nhn?code=52420'
driver.get(url)
soup = create_soup(url)
print(soup)
# 이름, 링크, 운용사, 수수료
etf_name = soup.select('#middle')


etf_link = f'https://finance.naver.com/item/coinfo.nhn?code='
company = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[1].text
commission = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[0].em.text.replace('%', '')
each_etf_info = [etf_name, etf_link, str(code), company, commission] # 1차 리스트 저장

# iframe으로 이동
driver.switch_to.frame('coinfo_cp')
time.sleep(1)
soup_iframe = BeautifulSoup(driver.page_source, "lxml")

# 시가총액
capital = soup_iframe.select_one('#status_grid_body').select('tr')[5].td.text.replace(',', '')[:-2]
each_etf_info.append(capital)

# 수익률
earnings = soup_iframe.select_one('#status_grid_body').select('tr')[-1].td.select('span')
for line in earnings:
    earning = line.text.replace('%', '').replace('+', '')
    each_etf_info.append(earning)

# 구성종목
top_10 = soup_iframe.select_one('#CU_grid_body').select('tr')[:10]
for line in top_10:
    sh_name = line.select('td')[0].text
    percent = line.select('td')[2].text
    if percent == '-':
        percent = ''
    else:
        percent = percent
    # 리스트 진짜 이방법밖에 없는거 실화..? 더 찾아보기
    each_etf_info.append(sh_name)  # 3차 리스트 저장
    each_etf_info.append(percent)
# csv에 입력
csv_writer.writerow(each_etf_info)

driver.quit()
print(f"ETF info completed")
print(time.strftime("%Y-%m-%d %H:%M:%S"))
# for code in etf_code:
#     time.sleep(5)
#     print(f"Start : {code}")
#     url = f'https://finance.naver.com/item/coinfo.nhn?code={code}'
#     driver.get(url)
#     soup = create_soup(url)
#
#     # 이름, 링크, 운용사, 수수료
#     etf_name = soup.select_one('#middle > div.h_company > div.wrap_company > h2 > a').text
#     print(etf_name)
#     etf_link = f'https://finance.naver.com/item/coinfo.nhn?code={code}'
#     company = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[1].text
#     commission = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[0].em.text.replace('%', '')
#     each_etf_info = [etf_name, etf_link, str(code), company, commission] # 1차 리스트 저장
#
#     # iframe으로 이동
#     driver.switch_to.frame('coinfo_cp')
#     time.sleep(1)
#     soup_iframe = BeautifulSoup(driver.page_source, "lxml")
#
#     # 시가총액
#     capital = soup_iframe.select_one('#status_grid_body').select('tr')[5].td.text.replace(',', '')[:-2]
#     each_etf_info.append(capital)
#
#     # 수익률
#     earnings = soup_iframe.select_one('#status_grid_body').select('tr')[-1].td.select('span')
#     for line in earnings:
#         earning = line.text.replace('%', '').replace('+', '')
#         each_etf_info.append(earning)
#
#     # 구성종목
#     top_10 = soup_iframe.select_one('#CU_grid_body').select('tr')[:10]
#     for line in top_10:
#         sh_name = line.select('td')[0].text
#         percent = line.select('td')[2].text
#         if percent == '-':
#             percent = ''
#         else:
#             percent = percent
#         # 리스트 진짜 이방법밖에 없는거 실화..? 더 찾아보기
#         each_etf_info.append(sh_name)  # 3차 리스트 저장
#         each_etf_info.append(percent)
#     # csv에 입력
#     csv_writer.writerow(each_etf_info)
#
# driver.quit()
# print(f"ETF info completed")
# print(time.strftime("%Y-%m-%d %H:%M:%S"))'''