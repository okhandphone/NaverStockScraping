import requests
from bs4 import BeautifulSoup
import os
import datetime
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side

# 엑셀 서식
user_font_bold = Font(name="맑은 고딕", size=11, bold=True)
user_font_red = Font(name="맑은 고딕", size=11, color="FE2E2E")
user_align = Alignment(horizontal="center", vertical="center")
orange_fill = PatternFill(patternType="solid", fgColor=Color("F3E2A9"))
gray_fill = PatternFill(patternType="solid", fgColor=Color("D8D8D8"))
user_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

# 뉴스 저장 경로 및 파일 만들기
news_path = "stock_data\\News"
if not os.path.exists(news_path):
    os.mkdir(news_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "데일리 뉴스"

    # 컬럼 너비 {열번호: 너비} # 키워드, 날짜, 타이틀, 요약
    width_dict = {"B": 12, "C": 11, "D": 49, "E": 65}
    for key, value in width_dict.items():
        ws.column_dimensions[f"{key}"].width = value

    # 헤더
    ws.append(['링크', '키워드', '날짜', '타이틀', '요약'])
    for row in ws["A1:A4"]:  # 행을 가져오고
        for cell in row:  # 셀을 가져온다
            cell.font = user_font_bold
            cell.alignment = user_align
            cell.fill = orange_fill

    wb.save(news_path + "\\데일리_뉴스.xlsx")
    print("새 엑셀 파일 생성")

# 이미 디렉토리가 있다면 엑셀 파일 로드
else:
    wb = load_workbook(news_path + "\\데일리_뉴스.xlsx")
    ws = wb.active

    print("데일리뉴스 엑셀파일 로드")

while True:
    keyword = input("키워드 입력 : ")

    # 끝이라고 입력하면 프로그램 종료
    if keyword == "끝" or keyword == "Rmx":
        break

    # # 이미지 폴더 만들기
    # img_path = f"stock_data\\News\\image"
    # if not os.path.exists(img_path):
    #     os.mkdir(img_path)

    # 오늘 날짜 함께 입력
    today = datetime.date.today()

    # 뉴스 페이지 '최신순' 첫번재 페이지
    url = f"https://search.naver.com/search.naver?where=news&query={keyword}&sm=tab_srt&sort=1&photo=0&field=0&reporter_article=&pd=0&ds=&de=&docid=&nso=so%3Add%2Cp%3Aall%2Ca%3Aall&mynews=0&refresh_start=0&related=0"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.72 Safari/537.36 Edg/89.0.774.45"}
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    t = soup.select('.news_area > a')
    img = soup.select('.news_wrap .dsc_thumb img')

    for i in range(len(t)):
        try:
            # 타이틀 및 링크
            title = t[i].text
            link = t[i]['href']

            ws.append([link, keyword, today, title])

            # 셀서식
            # for 문 안에서 엑셀을 컨트롤하면 매번저장하면서 중간에 에러가 나도 엑셀에 데이터는 담겨있음
            for row in ws["A:E"]:  # 행(튜플)
                for cell in row:  # 셀
                    cell.border = user_border

            for cell in ws["B"]: # 키워드
                cell.font = user_font_red
                cell.alignment = user_align
            ws["C1"].font = user_font_bold

            for cell in ws["C"]: # 날짜
                cell.alignment = user_align

            # 하이퍼링크
            for i in range(2, ws.max_row):
                ws[f"D{i}"].hyperlink = ws[f"A{i}"].value

            # 오토필터
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = 'A2'

            # 컬럼 숨기기 (URL)
            ws.column_dimensions["A"].hidden = True

            wb.save(news_path + "\\데일리_뉴스.xlsx")

            # # 이미지 다운 / 입력
            # img_url = img[i]['src']
            # response = requests.get(img_url)
            # response.raise_for_status()
            #
            # with open(img_path + f"\\{keyword}_{i + 1}.png", "wb") as f:
            #     f.write(response.content)
            #
            # img_for_excel = Image(img_path + f"\\{keyword}_{i + 1}.png")
            # insert_img = ws.add_image(img_for_excel, f"A{ws.max_row}")
            # ws.append([insert_img, link, keyword, today, title])
        except:
            pass