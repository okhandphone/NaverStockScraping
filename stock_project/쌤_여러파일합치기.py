
import os
import pandas as pd
from functools import reduce
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side

# # https://blog.naver.com/gisookhyun/221924898840
pd.options.display.max_columns = 50
pd.options.display.max_rows = 500

writer = pd.ExcelWriter("stock_data/데일리등락률_통합.xlsx")
for item in ["업종", "테마"]:
    '''
    1. 파일 블러오기
    '''
    print()
    print(f"{item} 데이터 시작")

    # 1-1 디렉토리에서 병합할 파일의 파일명 리스트로 불러오기
    dir_path = f"stock_data/{item}데일리등락률"
    file_list = [file for file in os.listdir(dir_path) if file != f"{item}데일리등락률_total.xlsx"]

    # 1-2 각 파일의 데이터를 df화 해서 리스트로 묶기
    df_list = []
    for file in file_list:
        df = pd.read_csv(dir_path+"/"+file, index_col=1) # 인덱스 업종 코드
        date = file.split("_")[2].replace(".csv", "")
        print(date, "데이터 개수 : ", len(df))
        # 컬럼명 바꾸기 :'전일대비 등락률' > date
        df.rename(columns={'전일대비 등락률': date}, inplace=True)  # 참고 : https://rfriend.tistory.com/468
        df_list.append(df)

    # 1-3 df_merged = df_list 하나의 데이터 프레임으로 병합
    df_merged = reduce(lambda left, right: pd.merge(left, right, on=[f'{item}코드', f'{item}명'], how='outer'), df_list)
    df_merged.drop_duplicates(f'{item}명', inplace=True)  # 중복값 제거
    df_merged['전체등락률'] = round(df_merged.iloc[:, 2:].sum(axis=1), 2)  # 마지막 컬럼에 전체 등락률 합산 삽입
    df_merged.sort_values(by='전체등락률', inplace=True) # 전체 등락률로 정렬

    # 1-4 엘셀로 저장
    df_merged.to_excel(writer, sheet_name=f"{item}")
writer.save()
print("판다스 완료")
#
# '''
# # 2. 엑셀 다듬기
# '''
# # 셀서식
# align_center = Alignment(horizontal="center", vertical="center")
# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
#                      bottom=Side(style="thin"))
#
# # 글꼴 서식
# font_normal = Font(name="맑은 고딕", size=11)
# font_bold = Font(name="맑은 고딕", size=11, bold=True)
#
# # color_fill
# orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
#
# light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
# pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
# pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
# pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
# dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
# red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))
#
# light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
# blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
# blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
# blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
# dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
# navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))
#
# # 워크북 불러오기
#
# wb = load_workbook("stock_data/데일리등락률_통합.xlsx")
# for item2 in ["업종", "테마"]:
#     ws = wb[f"{item2}"]
#
#     # 컬럼너비
#     ws.column_dimensions["B"].width = 30
#
#     # 헤더 서식
#     for cell in ws["A"]:
#         cell.font = font_normal
#     for y in range(1, ws.max_column + 1):
#         ws.cell(row=1, column=y).font = font_bold
#         ws.cell(row=1, column=y).alignment = align_center
#         ws.cell(row=1, column=y).fill = orange_fill
#
#     # 등락률
#     for x in range(2, ws.max_row+1):
#         for y in range(3, ws.max_column+1):
#             if ws.cell(row=x, column=y).value == None:
#                 continue
#             if ws.cell(row=x, column=y).value > 5:
#                 ws.cell(row=x, column=y).fill = red_fill
#             elif ws.cell(row=x, column=y).value > 4:
#                 ws.cell(row=x, column=y).fill = dark_pink_fill
#             elif ws.cell(row=x, column=y).value > 3:
#                 ws.cell(row=x, column=y).fill = pink_fill_3
#             elif ws.cell(row=x, column=y).value > 2:
#                 ws.cell(row=x, column=y).fill = pink_fill_2
#             elif ws.cell(row=x, column=y).value > 1:
#                 ws.cell(row=x, column=y).fill = pink_fill
#             elif ws.cell(row=x, column=y).value > 0:
#                 ws.cell(row=x, column=y).fill = light_pink_fill
#             elif ws.cell(row=x, column=y).value > -1:
#                 ws.cell(row=x, column=y).fill = light_blue_fill
#             elif ws.cell(row=x, column=y).value > -2:
#                 ws.cell(row=x, column=y).fill = blue_fill
#             elif ws.cell(row=x, column=y).value > -3:
#                 ws.cell(row=x, column=y).fill = blue_fill_2
#             elif ws.cell(row=x, column=y).value > -4:
#                 ws.cell(row=x, column=y).fill = dark_blue_fill
#             else:
#                 ws.cell(row=x, column=y).fill = navy_fill
#
#     # 전체 보더
#     for x in range(1, ws.max_row + 1):
#         for y in range(1, ws.max_column + 1):
#             ws.cell(row=x, column=y).border = thin_border
#
#     # 오토필터
#     ws.auto_filter.ref = ws.dimensions
#     ws.freeze_panes = 'C2'
#
#     # 하이퍼링크
#     hyper_dict = {"업종": "https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no=",
#                   "테마": "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no="}
#     for i in range(2, ws.max_row):
#         ws[f"B{i}"].hyperlink = hyper_dict[f"{item2}"] + str(ws[f"A{i}"].value)
#
# wb.save("stock_data/데일리등락률_통합.xlsx")
# print("엑셀 서식 완료")
#
