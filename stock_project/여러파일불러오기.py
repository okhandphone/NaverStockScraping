import os
import pandas as pd
from pandas import DataFrame
from functools import reduce
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side



#
# '''
# 1. 파일 블러오기
# '''
# # 디렉토리 경로
# dir_path = f"stock_data/실시간벨류에이션"
# # 디렉토리 내 파일 리스트로 만들고 반복문 돌림
# file_list = [file for file in os.listdir(dir_path) if file != "realtime_stock_value.xlsx"]
# print(file_list)
#
# df_list = []
# for file in file_list:
#     df = pd.read_csv(dir_path+"/"+file, index_col=1)
#     print(len(df))
#     # '전일대비 등락률'을 날짜로 바꾸기
#     date = file.split("_")[2].replace(".csv", "")
#     print(date)
#     df.rename(columns={'전일대비 등락률': date}, inplace=True)  # 참고 : https://rfriend.tistory.com/468
#     df_list.append(df)
# print(df_list)
#
# # 데이터프레임 병합 # 참고 : https://stackoverflow.com/questions/44327999/python-pandas-merge-multiple-dataframes
# df_merged = reduce(lambda left,right: pd.merge(left, right, on=[f'{item}코드', f'{item}명'], how='outer'), df_list)
# print(df_merged)
#
# with pd.ExcelWriter(dir_path + f"/{item}데일리등락률_total.xlsx", mode='w', engine='openpyxl') as writer:
#     df_merged.to_excel(writer, sheet_name=f"{item}")
# print("1. 자료 취합완료")
#
#
# '''
# # 2. 엑셀 다듬기
# '''
# # 셀서식
# align_center = Alignment(horizontal="center", vertical="center")
# thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
#                      bottom=Side(style="thin"))
#
# # 글꼴 서식
# font_normal = Font(name="맑은 고딕", size=11)
# font_bold = Font(name="맑은 고딕", size=11, bold=True)
#
# # color_fill
# orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
#
# light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
# pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
# pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
# pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
# dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
# red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))
#
# light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
# blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
# blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
# blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
# dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
# navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))
#
# # 워크북 불러오기
# for item2 in ["업종", "테마"]:
#
#     dir_path = f"stock_data/{item2}데일리등락률"
#     wb = load_workbook(dir_path + f"/{item2}데일리등락률_total.xlsx")
#     ws = wb[f"{item2}"]
#
#     # 컬럼너비
#     ws.column_dimensions["B"].width = 30
#
#     # 헤더 서식
#     for cell in ws["A"]:
#         cell.font = font_normal
#
#     for y in range(1, ws.max_column + 1):
#         ws.cell(row=1, column=y).font = font_bold
#         ws.cell(row=1, column=y).alignment = align_center
#         ws.cell(row=1, column=y).fill = orange_fill
#
#     # 등락률
#     for x in range(2, ws.max_row+1):
#         for y in range(3, ws.max_column+1):
#             if ws.cell(row=x, column=y).value == None:
#                 continue
#             if ws.cell(row=x, column=y).value > 5:
#                 ws.cell(row=x, column=y).fill = red_fill
#             elif ws.cell(row=x, column=y).value > 4:
#                 ws.cell(row=x, column=y).fill = dark_pink_fill
#             elif ws.cell(row=x, column=y).value > 3:
#                 ws.cell(row=x, column=y).fill = pink_fill_3
#             elif ws.cell(row=x, column=y).value > 2:
#                 ws.cell(row=x, column=y).fill = pink_fill_2
#             elif ws.cell(row=x, column=y).value > 1:
#                 ws.cell(row=x, column=y).fill = pink_fill
#             elif ws.cell(row=x, column=y).value > 0:
#                 ws.cell(row=x, column=y).fill = light_pink_fill
#             elif ws.cell(row=x, column=y).value > -1:
#                 ws.cell(row=x, column=y).fill = light_blue_fill
#             elif ws.cell(row=x, column=y).value > -2:
#                 ws.cell(row=x, column=y).fill = blue_fill
#             elif ws.cell(row=x, column=y).value > -3:
#                 ws.cell(row=x, column=y).fill = blue_fill_2
#             elif ws.cell(row=x, column=y).value > -4:
#                 ws.cell(row=x, column=y).fill = dark_blue_fill
#             else:
#                 ws.cell(row=x, column=y).fill = navy_fill
#
#     # 전체 보더
#     for x in range(1, ws.max_row + 1):
#         for y in range(1, ws.max_column + 1):
#             ws.cell(row=x, column=y).border = thin_border
#
#     # 오토필터
#     ws.auto_filter.ref = ws.dimensions
#     ws.freeze_panes = 'C2'
#
#     # 하이퍼링크
#     hyper_dict = {
#         "업종": "https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no=",
#         "테마": "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no="
#     }
#
#     for i in range(2, ws.max_row):
#         ws[f"B{i}"].hyperlink = hyper_dict[f"{item2}"] + str(ws[f"A{i}"].value)
#
#     wb.save(dir_path + f"/{item2}데일리등락률_total.xlsx")
#
# print("2. 엑셀 서식 완료")
#








# https://blog.naver.com/gisookhyun/221924898840
pd.options.display.max_columns = 50
pd.options.display.max_rows = 500

'''
1. 파일 블러오기
'''
# 디렉토리 경로
dir_path = "stock_data/업종데일리등락률"
# 디렉토리 내 파일 리스트로 만들고 반복문 돌림
file_list = [file for file in os.listdir(dir_path) if not file == '업종데일리등락률_total.csv']
# print(file_list)
'''
['market_fluctuation_2021-03-08.csv', 'market_fluctuation_2021-03-09.csv', 'market_fluctuation_2021-03-10.csv', 'market_fluctuation_2021-03-11.csv', 'market_fluctuation_2021-03-12.csv', 'market_fluctuation_2021-03-15.csv', 'market_fluctuation_2021-03-16.csv', 'market_fluctuation_2021-03-17.csv', 'market_fluctuation_2021-03-18.csv', 'market_fluctuation_2021-03-19.csv']
'''
# {df["마켓코드"]: {market_name : df["마켓명"], 날짜: df["등락률"], 날짜: df["등락률"],... }, df["마켓코드"]: {market_name : df["마켓명"], 날짜: df["등락률"]}, ...}
columns_list = ["업종명"] # 데이터프레임 인덱스
file_path = "stock_data//업종데일리등락률/업종데일리등락률_total.csv"

# try:
#     writer = pd.ExcelWriter(file_path, mode='a', engine='openpyxl')

fluc_dict = {}

for file in file_list:

    df = pd.read_csv(dir_path + "/" + file)
    date = file.split("_")[2].replace(".csv", "")
    columns_list.append(date)

    print(f"{date} 진행 중")
    # print(len(df.iloc[:, 0]))
    # csv 자료 불러오기

    for i in range(len(df.iloc[:, 0])):
        code = df.iloc[i, 1]
        name = df.iloc[i, 0]
        fluc = df.iloc[i, 2]
        print(code)

#         # 딕셔너리 생성
#         if code not in fluc_dict:
#             fluc_dict[code] = {"업종명": name, date: fluc}
#         else:
#             fluc_dict[code][date] = fluc
# print(fluc_dict)
        # total_df = pd.DataFrame(fluc_dict[code], columns=columns_list, index=code)
        # print(total_df)

# # print(fluc_dict)
# print(columns_list)
# #
# for key, value in fluc_dict.items():
#     total_df = pd.DataFrame(value, columns=columns_list, index=[key])
#     print(total_df)
#     total_df.to_csv(file_path, encoding='utf-8-sig', mode="a")

    # total_df.to_excel(file_path, encoding='utf-8-sig')
# #
# total_df = pd.DataFrame(fluc_dict, columns=columns_list, index=list(fluc_dict.keys()))
# total_df.to_excel(file_path, encoding='utf-8-sig')
