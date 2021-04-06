
'''
엑셀 각종 서식 모음 파일
'''


import os
import csv

import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
import pandas as pd

'''
데이터 프레임으로 csv 파일 엑셀에 붙여넣기
'''
file_name = dir_path + "/realtime_stock_value.xlsx"
df = pd.read_csv(dir_path + f'/realtime_stock_value_{datetime.date.today()}.csv')

# pd.ExcelWriter() : 기존 파일 지워짐
if not os.path.exists(file_name):
    with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f"{datetime.date.today()}",index=False)
else:
    with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=f"{datetime.date.today()}",index=False)

print("df to 엑셀")

'''
엑셀불러오기
'''
# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ..
    print()
'''
엑셀 각종 서식 모음
'''

# 셀서식
align_center = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                     bottom=Side(style="thin"))

# 글꼴 서식
font_normal = Font(name="맑은 고딕", size=11)
font_bold = Font(name="맑은 고딕", size=11, bold=True)

# color_fill
orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))

light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))

light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))

# 워크북 불러오기
wb = load_workbook(file_name)
ws = wb[f"{datetime.date.today()}"]

# 컬럼너비 (종목명, 업종명, 상승종목 수, 하락종목 수)
width_dict = {"C": 21, "E": 19, "O": 12, "P": 12}
for key, value in width_dict:
    ws.column_dimensions[f"{key}"].width = value

# 헤더 서식
for row in ws["A1:M1"]:
    for cell in row:
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = orange_fill
        cell.border = thin_border

# 등락률 글자색
count_inc = 0 # 상승 종목 갯수
count_dec = 0 # 하락 종목 갯수
for col in ws[f"F2:F{ws.max_row}"]:
    for cell in col:
        if cell.value > 0:
            cell.font = font_red
            count_inc += 1
        else:
            cell.font = font_blue
            count_dec += 1

# 하이퍼링크
for i in range(2, ws.max_row):
    ws[f"C{i}"].hyperlink = ws[f"A{i}"].value
    print(ws[f"A{i}"].value)

# 상승 하락 종목 셀 서식
ws["O1"].value = "상승 종목 수"
ws["P1"].value = "하락 종목 수"
ws["O1"].fill = dark_pink_fill
ws["P1"].fill = dark_blue_fill
ws["O2"].value = count_inc
ws["P2"].value = count_dec

for row in ws[f"O1:P1"]:
    for cell in row:
        cell.font = font_bold

for row in ws[f"O1:P2"]:
    for cell in row:
        cell.border = thin_border
        cell.alignment = align_center

print(count_inc, count_dec)

# 전체 보더
for row in ws[f"A2:M{ws.max_row}"]:
    for cell in row:
        cell.border = thin_border

# 오토필터
ws.auto_filter.ref = ws.dimensions
ws.freeze_panes = 'A2'

# 컬럼 숨기기 (URL)
ws.column_dimensions["A"].hidden= True

wb.save(file_name)
print("엑셀 서식 완료")

# 셀서식
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))

    # 글꼴 서식
    font_normal = Font(name="맑은 고딕", size=11)
    font_bold = Font(name="맑은 고딕", size=11, bold=True)

    # color_fill
    orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))

    light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
    pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
    pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
    pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
    dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
    red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))

    light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
    blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
    blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
    blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
    dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
    navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))

    # 워크북 불러오기
    wb = load_workbook("stock_data/데일리등락률_통합.xlsx")
    for item2 in ["업종", "테마"]:
        ws = wb[f"{item2}"]

        # 컬럼너비
        ws.column_dimensions["B"].width = 30

        # 헤더 서식
        for cell in ws["A"]:
            cell.font = font_normal
        for y in range(1, ws.max_column + 1):
            ws.cell(row=1, column=y).font = font_bold
            ws.cell(row=1, column=y).alignment = align_center
            ws.cell(row=1, column=y).fill = orange_fill

        # 등락률
        for x in range(2, ws.max_row + 1):
            for y in range(3, ws.max_column + 1):
                if ws.cell(row=x, column=y).value == None:
                    continue
                if ws.cell(row=x, column=y).value > 5:
                    ws.cell(row=x, column=y).fill = red_fill
                elif ws.cell(row=x, column=y).value > 4:
                    ws.cell(row=x, column=y).fill = dark_pink_fill
                elif ws.cell(row=x, column=y).value > 3:
                    ws.cell(row=x, column=y).fill = pink_fill_3
                elif ws.cell(row=x, column=y).value > 2:
                    ws.cell(row=x, column=y).fill = pink_fill_2
                elif ws.cell(row=x, column=y).value > 1:
                    ws.cell(row=x, column=y).fill = pink_fill
                elif ws.cell(row=x, column=y).value > 0:
                    ws.cell(row=x, column=y).fill = light_pink_fill
                elif ws.cell(row=x, column=y).value > -1:
                    ws.cell(row=x, column=y).fill = light_blue_fill
                elif ws.cell(row=x, column=y).value > -2:
                    ws.cell(row=x, column=y).fill = blue_fill
                elif ws.cell(row=x, column=y).value > -3:
                    ws.cell(row=x, column=y).fill = blue_fill_2
                elif ws.cell(row=x, column=y).value > -4:
                    ws.cell(row=x, column=y).fill = dark_blue_fill
                else:
                    ws.cell(row=x, column=y).fill = navy_fill

        # 전체 보더
        for x in range(1, ws.max_row + 1):
            for y in range(1, ws.max_column + 1):
                ws.cell(row=x, column=y).border = thin_border

        # 오토필터
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'C2'

        # 하이퍼링크
        hyper_dict = {"업종": "https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no=",
                      "테마": "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no="}
        for i in range(2, ws.max_row):
            ws[f"B{i}"].hyperlink = hyper_dict[f"{item2}"] + str(ws[f"A{i}"].value)

    wb.save("stock_data/데일리등락률_통합.xlsx")
    print("엑셀 서식 완료")