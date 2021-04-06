# This Python file uses the following encoding: utf-8
import os, sys


#  User-Agent
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36"}

'''
1. 업종 종목 코드
'''

# 0. 한국거래소 종목, 업종 코드
# 엑셀 파일 아닌 홈페이지에서 바로 받는 법 찾아보기
def get_share_code_from_KRX():
    # openpyxl로 엑셀 파일 읽기
    from openpyxl import load_workbook  # 파일 불러오기
    wb = load_workbook("stock_data\\krxcode.xlsx")  # 불러올 파일명 넣기
    ws = wb.active

    print("get Share Code Dict")
    code_list = []
    code_dict = {}
    for x in range(2, ws.max_row + 1):
        code = [ws.cell(row=x, column=y).value for y in range(1, ws.max_column + 1)]
        code_list.append(code)
        code_dict.setdefault(f"{code[0]}", {"Share Name": code[1], "Market Code": str(code[2]), "Market Name": code[3]})

    print("get Share Code Done")
    return code_dict

# 1. 네이버에서 업종코드 받기
def get_market_code_from_naver():
    from scraping.web_scraping import create_soup
    soup = create_soup('https://finance.naver.com/sise/sise_group.nhn?type=upjong')
    table = soup.select('table.type_1 > tr > td > a')

    # 기타 항목 제거 위해 딕셔너리로 반환
    mk_code_dict = {}
    for line in table:
        market_name = line.text
        market_code = line['href'].replace("/sise/sise_group_detail.nhn?type=upjong&no=", "")
        mk_code_dict[market_name] = market_code
    del mk_code_dict['기타']
    print(mk_code_dict)
    print('mk_code_dict completed')
    # mk_code_dict를 리턴 {업종명: 업종코드}
    return mk_code_dict

# 2. 네이버 업종코드를 통해 종목 코드 받기
def get_share_code_from_naver():
    from scraping.web_scraping import create_soup
    from openpyxl import Workbook
    import time

    wb = Workbook()
    ws = wb.active
    ws.title = "네이버 업종 종목코드"
    ws.append(['종목코드', '종목명', '업종코드', '업종명'])

    # 업종코드 딕셔너리
    mk_code_dict = get_market_code_from_naver()

    # key = 업종명, value = 업종코드
    for key, value in mk_code_dict.items():
        print(f"Start {value} Category")
        # time.sleep(2)
        url = f'https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no={value}'
        soup = create_soup(url)
        table = soup.select('#contentarea > div.box_type_l')[1].select('tbody > tr > td.name')
        for line in table:
            share_name = line.text
            share_code = line.a['href'].replace("/item/main.nhn?code=", "")
            # 마켓과 종목 코드를 엑셀파일로 저장
            ws.append([share_code, share_name, value, key])
    wb.save("stock_data\\종목코드.xlsx")
    print("Total Code file Completed")

# 3. 엑셀에서 업종코드 딕셔너리로 가져오기
def get_code_from_excel():
    from openpyxl import load_workbook  # 파일 불러오기

    wb = load_workbook("stock_data\\종목코드.xlsx")
    ws = wb.active

    code_dict = {}
    for x in range(2, ws.max_row + 1):
        code = [ws.cell(row=x, column=y).value for y in range(1, ws.max_column + 1)]
        code_dict.setdefault(f"{code[0]}", {"Share Name": code[1], "Market Code": str(code[2]), "Market Name": code[3]})
    print(code_dict)
    print("code_dict completed")
    # 종합 코드딕트 생성 {종콕코드 : {"Share Name": 종목명, "Market Code": 업종코드, "Market Name": 업종명}
    return code_dict


'''
2. 네이버 테마 정보
'''

#  4. 네이버 테마 코드 받기
def get_theme_code_from_naver():
    import time
    from scraping.web_scraping import create_soup
    theme_code_list = [] # 테마 코드 리스트로 저장
    print("Start get share code")
    for page in range(1, 7):
        print(f"On page : {page}")
        url = f"https://finance.naver.com/sise/theme.nhn?&page={page}"
        soup = create_soup(url)

        theme_data = soup.select_one("table.type_1").select("td.col_type1")
        for line in theme_data:
            theme_code = line.select_one("a")["href"].replace("/sise/sise_group_detail.nhn?type=theme&no=", "")
            theme_code_list.append(theme_code)
    print("Theme code done")
    # 테마 코드 리스트를 리턴
    return theme_code_list


#  5. 네이버 테마별 테마명, 테마코드, 종목명, 종목코드, 설명 등 추출
def get_theme_share_info():
    import csv
    import time
    import datetime
    from scraping.web_scraping import create_soup

    print("Start get share code")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))
    csv_file = open(f"stock_data\\theme_share_info_{datetime.date.today()}.csv", "w", encoding='utf-8-sig', newline="")
    csv_writer = csv.writer(csv_file)

    # 헤더
    header = (['테마명', '테마코드', '종목명', '종목코드', '종목정보', '종목링크'])
    csv_writer.writerow(header)

    # 테마 코드 리스트
    th_code_list = get_theme_code_from_naver()

    # th_sh_info = [] # 테마명, 테마코드, 테마설명, 종목명, 종목코드, 종목 설명 리스트화
    for th_code in th_code_list:
        print(f"theme code: {th_code}")
        # time.sleep(3)
        url = f'https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no={th_code}'
        soup = create_soup(url)
        # 테마명, 테마 설명 가져오기
        th_name = soup.select_one('#contentarea_left > table > tr > td > div > div > strong').get_text()
        th_info = soup.select('#contentarea_left > table > tr > td')[1].p.get_text()
        # th_sh_info.append([th_name, th_code, '', '', th_info]) # 리스트에 추가
        # ws.append([th_name, th_code, '', '', th_info]) # 엑셀에 저장
        csv_writer.writerow([th_name, str(th_code), '', '', th_info])

        # 테마별 종목명, 종목코드
        sh_data = soup.find('tbody').select('tr')[:-2]
        for tag in sh_data:
            # sh_link = "https://finance.naver.com/" + tag.select('td')[0].a['href']
            sh_name = tag.select('td')[0].text
            sh_code = tag.select('td')[0].a['href'].replace("/item/main.nhn?code=", "")
            sh_info = tag.select('td')[1].p.text
            sh_link = "https://finance.naver.com/item/main.nhn?code=" + sh_code
            # th_sh_info.append([th_name, th_code, sh_name, sh_code, sh_info, sh_link])  # 리스트에 추가
            # ws.append([th_name, th_code, sh_name, sh_code, sh_info, sh_link]) # 엑셀에 저장
            csv_writer.writerow([th_name, str(th_code), sh_name, str(sh_code), sh_info, sh_link])

    csv_file.close()
    print("Share code done")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))


'''
3. 실적
'''
def get_growth_rate(data_list, i_list):
    gr_list = []
    for i in i_list:
        latter = data_list[i]
        pre = data_list[i - 1]
        if latter == '' or pre == '': gr_list.append('')
        elif latter == 0 and pre == 0: gr_list.append(0)
        elif pre == 0:
            pre = 1
            gr_list.append(round((latter - pre) / abs(pre) * 100, 2))
        else:
            gr_list.append(round((latter - pre) / abs(pre) * 100, 2))

    return gr_list

# 기간별 실적 스크래핑
def get_financial_info():
    import time
    import datetime
    from bs4 import BeautifulSoup
    from pandas import DataFrame
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import pandas as pd

    pd.options.display.max_columns = 50
    pd.options.display.max_rows = 500

    # headless
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument("window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
    driver = webdriver.Chrome(options=options)

    # 종목코드 딕셔너리
    code_dict = get_code_from_excel()
    # header = 기간,종목코드,종목명,업종코드,업종명,매출액,salesVar,영업이익,opVar,영업이익(발표기준),세전계속사업이익,당기순이익,npVar,당기순이익(지배),당기순이익(비지배),자산총계,부채총계,자본총계,자본총계(지배),자본총계(비지배),자본금,영업활동현금흐름,투자활동현금흐름,재무활동현금흐름,CAPEX,FCF,이자발생부채,영업이익률,순이익률,ROE(%),ROA(%),부채비율,자본유보율,EPS(원),PER(배),BPS(원),PBR(배),현금DPS(원),현금배당수익률,현금배당성향(%),발행주식수(보통주)

    print("재무정보 크롤링 시작")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))
    # code_dict = {'016610': {'Share Name': 'DB금융투자 ', 'Market Code': '12', 'Market Name': '증권'},
    #              '005830': {'Share Name': 'DB손해보험 ', 'Market Code': '190', 'Market Name': '손해보험'},
    #              '000990': {'Share Name': 'DB하이텍 ', 'Market Code': '202', 'Market Name': '반도체와반도체장비'},
    #              '139130': {'Share Name': 'DGB금융지주 ', 'Market Code': '20', 'Market Name': '은행'},
    #              '068790': {'Share Name': 'DMS *', 'Market Code': '199', 'Market Name': '디스플레이장비및부품'},
    #              '034220': {'Share Name': 'LG디스플레이 ', 'Market Code': '222', 'Market Name': '디스플레이패널'},
    #              '032820': {'Share Name': '우리기술 *', 'Market Code': '197', 'Market Name': '전자장비와기기'},
    #              '066790': {'Share Name': '씨씨에스 *', 'Market Code': '204', 'Market Name': '방송과엔터테인먼트'},
    #              '148250': {'Share Name': '알엔투테크놀로지 *', 'Market Code': '136', 'Market Name': '통신장비'}}

    # iframe 직접 접근 # 빨라짐
    for key, value in code_dict.items():
        try:
            driver.get(f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={key}')
            print(f"Start: {key}")
            time.sleep(3)
            driver.implicitly_wait(3)
            wait = WebDriverWait(driver, 3)

            # 연간 분기 돌아가면서 정보 받아오기
            period_dict = {'cns_Tab21': '연간', 'cns_Tab22': '분기'}
            for p_key, p_val in period_dict.items():
                # print(f"Start: {key}, {p_val} 정보")
                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#{p_key}'))).click()
                time.sleep(1)

                # iframe 에서 html정보 추출
                soup = BeautifulSoup(driver.page_source, "html.parser")

                # 기간 정보
                period_data = soup.select('table.gHead01')[3].select('thead > tr > th')[2:]
                period = []
                # 정규식으로 간단히 지울 방법 생각해보기
                for line in period_data:
                    if line.span.text == "(IFRS연결)":
                        date = line.text.replace("(IFRS연결)", "").strip()
                    elif line.span.text == "(IFRS별도)":
                        date = line.text.replace("(IFRS별도)", "").strip()
                    else:
                        date = line.text.replace("(GAAP개별)", "").strip()
                    period.append(date)

                # financial_dict에 코드정보 입력
                financial_dict = {'Share Code': f'{str(key)}'}
                financial_dict.update(value)  # code_dict의 밸류값 추가

                # 재무정보
                financial_table = soup.select('table.gHead01')[3].select('tbody > tr')
                for line in financial_table:
                    name = line.select_one('th').text.strip()  # 재무 항목명
                    data = line.select('td')  # 재무 데이터
                    financial_dict.setdefault(name, [num.text.replace(",", "") for num in data])  # 딕셔너리에 추가
                    #  매출액, 영업이익, 당기순이익의 성장률 데이터 추가
                    gr_dict = {'매출액': 'salesVar', '영업이익': 'opVar', '당기순이익': 'npVar'}
                    if name in gr_dict:
                        data_list = list(map(lambda a: '' if a == '' else float(a), financial_dict[name]))  # float으로 형변환
                        i_list = list(range(1, len(data_list)))
                        gr_list = get_growth_rate(data_list, i_list)
                        gr_list.insert(0, '') # 성장률 맨 앞에 빈 자리 넣어주기
                        financial_dict.setdefault(gr_dict[name], gr_list)
                # print(financial_dict)

                # financial_dict를 데이터 프레임으로
                financial_df = DataFrame(financial_dict, columns=list(financial_dict.keys()), index=period)
                financial_df.to_csv(f"stock_data/재무실적/{p_val}실적_raw_{datetime.date.today()}.csv", mode="a", header=False, encoding='utf-8-sig')

        except:
            pass
            print(f'pass: {key}')
    print("재무 크롤링 완료")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))
    print()
    driver.quit()


# 재무자료 합치기 # 테스트 해야함
def concat_financial_info(old_file_path): # 재무실적_통합

    import datetime
    import pandas as pd

    pd.options.display.max_columns = 50
    pd.options.display.max_rows = 500

    header = ['기간', '종목코드', '종목명', '업종코드', '업종명', '매출액', 'salesVar', '영업이익', 'opVar', '영업이익(발표기준)', '세전계속사업이익', '당기순이익', 'npVar', '당기순이익(지배)', '당기순이익(비지배)',
              '자산총계', '부채총계', '자본총계', '자본총계(지배)', '자본총계(비지배)', '자본금', '영업활동현금흐름', '투자활동현금흐름', '재무활동현금흐름', 'CAPEX', 'FCF', '이자발생부채', '영업이익률', '순이익률',
              'ROE(%)', 'ROA(%)', '부채비율', '자본유보율', 'EPS(원)', 'PER(배)', 'BPS(원)', 'PBR(배)', '현금DPS(원)', '현금배당수익률', '현금배당성향(%)', '발행주식수(보통주)']

    # writer = pd.ExcelWriter(f"stock_data/재무실적/종목실적_raw_{datetime.date.today()}.xlsx")
    for period in ['연간', '분기']:

        # 기간별 재무 정보 불러오기
        print(f'{period} 시작')
        df = pd.read_csv(f"stock_data/재무실적/로우데이터/{period}실적_raw_2021-03-31.csv", names=header)  # new 재무자료
        print(df.head(10))
        old_df = pd.read_excel(old_file_path, dtype={'종목코드': str}, sheet_name=f'{period}', index_col=None)  # pre 재무자료
        print(old_df.head(10))
        # print(df.head(50))

        concat_df = pd.concat([df, old_df])  # 이전 재무 데이터와 합치기
        concat_df.dropna(subset=['당기순이익'], inplace=True)  # 당기순이익이 비어있는 행 제거
        concat_df.drop_duplicates(['기간', '종목코드'], inplace=True)  # 중복 제거
        concat_df.sort_values(by=['종목명', '기간'], inplace=True)  # 종목명과 기간으로 정렬
        print(concat_df.head(50))

        # 엑셀로 저장
        concat_df.to_excel(f"stock_data/재무실적/로우데이터/재무실적_통합_{datetime.date.today()}.xlsx", index=None)
        print(f'{period} 실적 엑셀 저장 완료')

# 실시간 밸류 확인 (네이버 업종)
def get_realtime_value():
    import os
    import csv
    import time
    import datetime
    from bs4 import BeautifulSoup
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
    import pandas as pd

    '''
    # 1. CSV 파일 생성
    '''
    dir_path = "stock_data/실시간벨류에이션"
    if not os.path.exists(dir_path): os.mkdir(dir_path)
    csv_file = open(dir_path + f'\\realtime_stock_value.csv', 'w', encoding='utf-8-sig',newline='')
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(['URL', '종목코드', '종목명', '업종코드', '업종명', '등락률', '시가총액 (억 원)', 'PER', 'ROE', 'PEG', 'ROA', 'PBR', '유보율'])

    '''
    # 2. 네이버 업종 크롤링
    '''
    # headless
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument("window-size=1920x1080")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
    driver = webdriver.Chrome(options=options)

    # 마켓코드 받아오기
    mk_code_dict = get_market_code_from_naver()
    print()
    print("실시간 밸류 스크랩 시작")
    for mk_name, mk_code in mk_code_dict.items():
        # try:
        # time.sleep(3)
        print()
        print(f"Start {mk_name} Category")
        url = f'https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no={mk_code}'

        # 셀레니움으로 받아야 옵션 정보가 유지됨
        driver.get(url)
        driver.implicitly_wait(3)
        wait = WebDriverWait(driver, 3)

        titles = driver.find_elements_by_css_selector("#contentarea > div:nth-child(5) > table > thead > tr:nth-child(1) > th")
        title = [line.text for line in titles]
        print(title, "sel")

        # title에 '거래량'이 있을 경우 옵션변경
        if '거래량' in title:

            # 기존 옵션 제거 : 거래량, 매수호가, 거래대금, 매도호가, 전일거래량
            remove_list = [1, 2, 3, 8, 9]
            for num in remove_list:
                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#option{num}'))).click()

            # 원하는 옵션 클릭 : 시가총액, PER, ROE, ROA, PBR, 유보율
            remove_list = [4, 6, 12, 18, 24, 27]
            for num in remove_list:
                wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f'#option{num}'))).click()

            # 옵션 적용 클릭
            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.item_btn > a'))).click()  # 적용하기
            time.sleep(1)

        # html 파싱
        soup = BeautifulSoup(driver.page_source, "lxml")
        # 타이틀 다시 확인

        titles = soup.select("div.box_type_l th")
        title = [line.text for line in titles]
        print(title, "soup")
        table = soup.select('#contentarea > div.box_type_l')[1].select('tbody > tr')[:-2]

        # 종목별 정보 추출 및 기록
        for line in table:
            # 리스트에 이름 및 코드 추가
            share_name = line.td.text
            share_code = line.td.a['href'].replace("/item/main.nhn?code=", "")
            share_link = "https://finance.naver.com/item/coinfo.nhn?code=" + share_code
            info_list = [share_link, str(share_code), share_name, str(mk_code), mk_name]

            # 벨류 데이터 추가 (등락률 부터)
            data = line.select('td')[3:-1]
            for num in data:
                num = num.text.replace(',', '').replace('+', '').replace('%', '')
                if num == '':
                    num = ''
                else:
                    num = float(num)
                info_list.append(num)

            # PEG 밸류 추가
            per = info_list[7]
            roe = info_list[8]
            # per, roe 값을 기준으로 info_list 값 달라짐
            if per == '' or roe == '':  # 값이 없는 경우
                info_list.insert(9, '')
            elif per > 0 and roe > 0:  # 0보다 크거나 같으면 peg 계산
                peg = per / roe
                info_list.insert(9, f"{peg:.1f}")
            elif per <= 0 or roe <= 0:
                info_list.insert(9, '')  # 마이너스인 경우 '' 반환

            # csv에 기록
            csv_writer.writerow(info_list)

    driver.quit()
    csv_file.close()
    print("실시간 밸류 크롤링 완료")

    # '''
    # # 3. realtime_stock_value.xlsx 파일에 날짜 이름 시트로 붙여넣기
    # '''
    # # 아예 엑셀파일로 바로 생성하는 것 생각해보기
    #
    # file_name = dir_path + "/realtime_stock_value.xlsx"
    # df = pd.read_csv(dir_path + f'/realtime_stock_value_{datetime.date.today()}.csv')
    #
    # if not os.path.exists(file_name):
    #     with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
    #         df.to_excel(writer, sheet_name=f"{datetime.date.today()}", index=False)
    # else:
    #     with pd.ExcelWriter(file_name, mode='a', engine='openpyxl') as writer:
    #         df.to_excel(writer, sheet_name=f"{datetime.date.today()}", index=False)
    # print("2. df to 엑셀")
    #
    #
    # '''
    # # 4. 엑셀 서식 추가
    # '''
    # # 셀서식
    # align_center = Alignment(horizontal="center", vertical="center")
    # thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
    #                      bottom=Side(style="thin"))
    #
    # # 글꼴 서식
    # font_bold = Font(name="맑은 고딕", size=11, bold=True)
    # font_red = Font(name="맑은 고딕", size=11, color="FE2E2E")
    # font_blue = Font(name="맑은 고딕", size=11, color="1E88E5")
    #
    # # color_fill
    # orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
    # dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("EC7063"))
    # dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("5DADE2"))
    #
    # # 워크북 불러오기
    # wb = load_workbook(file_name)
    # ws = wb[f"{datetime.date.today()}"]
    #
    # # 컬럼너비 (종목명, 업종명, 상승종목 수, 하락종목 수)
    # width_dict = {"C": 21, "E": 19, "O": 13, "P": 13, "Q": 13}
    # for key, value in width_dict.items():
    #     ws.column_dimensions[f"{key}"].width = value
    #
    # # 헤더 서식
    # for row in ws["A1:M1"]:
    #     for cell in row:
    #         cell.font = font_bold
    #         cell.alignment = align_center
    #         cell.fill = orange_fill
    #         cell.border = thin_border
    #
    # # 등락률 글자색
    # count_inc = 0  # 상승 종목 갯수
    # count_dec = 0  # 하락 종목 갯수
    # for col in ws[f"F2:F{ws.max_row}"]:
    #     for cell in col:
    #         if cell.value > 0:
    #             cell.font = font_red
    #             count_inc += 1
    #         else:
    #             cell.font = font_blue
    #             count_dec += 1
    #
    # # 하이퍼링크
    # for i in range(2, ws.max_row):
    #     ws[f"C{i}"].hyperlink = ws[f"A{i}"].value
    #
    # # 상승 하락 종목 셀 서식
    # ws["O1"].value = "상승 종목 수"
    # ws["P1"].value = "하락 종목 수"
    # ws["Q1"].value = "총 종목 수"
    # ws["O1"].fill = dark_pink_fill
    # ws["P1"].fill = dark_blue_fill
    # ws["Q1"].fill = orange_fill
    # ws["O2"].value = count_inc
    # ws["P2"].value = count_dec
    # ws["Q2"].value = count_inc + count_dec
    #
    # for row in ws[f"O1:Q1"]:
    #     for cell in row:
    #         cell.font = font_bold
    #
    # for row in ws[f"O1:Q2"]:
    #     for cell in row:
    #         cell.border = thin_border
    #         cell.alignment = align_center
    #
    # print(count_inc, count_dec, count_inc + count_dec)
    #
    # # 전체 보더
    # for row in ws[f"A2:M{ws.max_row}"]:
    #     for cell in row:
    #         cell.border = thin_border
    #
    # # 오토필터
    # ws.auto_filter.ref = ws.dimensions
    # ws.freeze_panes = 'A2'
    #
    # # 컬럼 숨기기 (URL)
    # ws.column_dimensions["A"].hidden = True
    #
    # wb.save(file_name)
    # print("3. 엑셀 서식 완료")
    # print("실시간 밸류 완료")
    # print()



'''
4. 수급
'''

def get_market_fluctuation():
    import csv
    import datetime
    from scraping.web_scraping import create_soup

    csv_file = open(f"stock_data\\업종데일리등락률\\market_fluctuation_{datetime.date.today()}.csv", "w", encoding='utf-8-sig', newline="")
    csv_writer = csv.writer(csv_file)

    # period = datetime.date.today()

    # 헤더
    header = (['업종명', '업종코드', '전일대비 등락률'])
    csv_writer.writerow(header)

    # 종목코드 종목명 전일대비등락률
    print("데일리 업종별 등락률 스크랩 시작")

    # 네이버 금융 업종
    url = "https://finance.naver.com/sise/sise_group.nhn?type=upjong"
    soup = create_soup(url)
    theme_data = soup.select_one("table.type_1").select("tr")[3:]
    for line in theme_data:
        td = line.select('td')
        if len(td) <= 1:  # 의미없는 자료 제거
            continue
        market_name = td[0].text.strip()
        market_code = td[0].a["href"].replace("/sise/sise_group_detail.nhn?type=upjong&no=", "")
        fluc = td[1].text.strip().replace('%', '').replace('+', '')
        # link = "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no=" + market_code
        csv_writer.writerow([market_name, market_code, fluc])
    print("업종별 등락률 완료")
    print()


def get_theme_fluctuation():
    import csv
    import time
    import datetime
    from scraping.web_scraping import create_soup

    csv_file = open(f"stock_data\\테마데일리등락률\\theme_fluctuation_{datetime.date.today()}.csv", "w", encoding='utf-8-sig', newline="")
    csv_writer = csv.writer(csv_file)

    # period = datetime.date.today()

    # 헤더
    header = (['테마명', '테마코드', '전일대비 등락률'])
    csv_writer.writerow(header)

    # 테마명 테마코드 전일대비등락률
    print("데일리 테마별 등락률 스크랩 시작")

    for page in range(1, 7):
        print(f"On page : {page}")
        # 네이버 금융 테마별 시세
        url = f"https://finance.naver.com/sise/theme.nhn?&page={page}"
        time.sleep(1.5)

        soup = create_soup(url)
        theme_data = soup.select("td.col_type1")
        flucs = soup.select("td.col_type2")
        for i in range(len(theme_data)):
            theme_name = theme_data[i].text.strip()
            theme_code = theme_data[i].a["href"].replace("/sise/sise_group_detail.nhn?type=theme&no=", "")
            fluc = flucs[i].text.strip().replace("+", "").replace("%", "")
            print(theme_name, theme_code, fluc)
            print(i)
            csv_writer.writerow([theme_name, theme_code, fluc])
    print("테마별 등락률 완료")
    print()

'''
5. 파일 통합
'''
def compile_fluc_files():
    import os
    import pandas as pd
    from functools import reduce
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side

    # # https://blog.naver.com/gisookhyun/221924898840
    pd.options.display.max_columns = 50
    pd.options.display.max_rows = 500

    writer = pd.ExcelWriter("stock_data/데일리등락률_통합.xlsx")
    for item in ["업종", "테마"]:
        '''
        1. 파일 블러오기
        '''
        print()
        print(f"{item} 데이터 시작")

        # 1-1 디렉토리에서 병합할 파일의 파일명 리스트로 불러오기
        dir_path = f"stock_data/{item}데일리등락률"
        file_list = [file for file in os.listdir(dir_path) if file != f"{item}데일리등락률_total.xlsx"]

        # 1-2 각 파일의 데이터를 df화 해서 리스트로 묶기
        df_list = []
        for file in file_list:
            df = pd.read_csv(dir_path + "/" + file, index_col=1)  # 인덱스 업종 코드
            date = file.split("_")[2].replace(".csv", "")
            print(date, "데이터 개수 : ", len(df))
            # 컬럼명 바꾸기 :'전일대비 등락률' > date
            df.rename(columns={'전일대비 등락률': date}, inplace=True)  # 참고 : https://rfriend.tistory.com/468
            df_list.append(df)

        # 1-3 df_merged = df_list 하나의 데이터 프레임으로 병합
        df_merged = reduce(lambda left, right: pd.merge(left, right, on=[f'{item}코드', f'{item}명'], how='outer'),
                           df_list)
        df_merged.drop_duplicates(f'{item}명', inplace=True)  # 중복값 제거
        df_merged['전체등락률'] = round(df_merged.iloc[:, 2:].sum(axis=1), 2)  # 마지막 컬럼에 전체 등락률 합산 삽입
        df_merged.sort_values(by='전체등락률', inplace=True)  # 전체 등락률로 정렬

        # 1-4 엘셀로 저장
        df_merged.to_excel(writer, sheet_name=f"{item}")
    writer.save()
    print("등락률 데이터 통합 완료")

    # '''
    # # 2. 엑셀 다듬기
    # '''
    # # 셀서식
    # align_center = Alignment(horizontal="center", vertical="center")
    # thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
    #                      bottom=Side(style="thin"))
    #
    # # 글꼴 서식
    # font_normal = Font(name="맑은 고딕", size=11)
    # font_bold = Font(name="맑은 고딕", size=11, bold=True)
    #
    # # color_fill
    # orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
    #
    # light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
    # pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
    # pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
    # pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
    # dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
    # red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))
    #
    # light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
    # blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
    # blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
    # blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
    # dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
    # navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))
    #
    # # 워크북 불러오기
    # wb = load_workbook("stock_data/데일리등락률_통합.xlsx")
    # for item2 in ["업종", "테마"]:
    #     ws = wb[f"{item2}"]
    #
    #     # 컬럼너비
    #     ws.column_dimensions["B"].width = 30
    #
    #     # 헤더 서식
    #     for cell in ws["A"]:
    #         cell.font = font_normal
    #     for y in range(1, ws.max_column + 1):
    #         ws.cell(row=1, column=y).font = font_bold
    #         ws.cell(row=1, column=y).alignment = align_center
    #         ws.cell(row=1, column=y).fill = orange_fill
    #
    #     # 등락률
    #     for x in range(2, ws.max_row + 1):
    #         for y in range(3, ws.max_column + 1):
    #             if ws.cell(row=x, column=y).value == None:
    #                 continue
    #             if ws.cell(row=x, column=y).value > 5:
    #                 ws.cell(row=x, column=y).fill = red_fill
    #             elif ws.cell(row=x, column=y).value > 4:
    #                 ws.cell(row=x, column=y).fill = dark_pink_fill
    #             elif ws.cell(row=x, column=y).value > 3:
    #                 ws.cell(row=x, column=y).fill = pink_fill_3
    #             elif ws.cell(row=x, column=y).value > 2:
    #                 ws.cell(row=x, column=y).fill = pink_fill_2
    #             elif ws.cell(row=x, column=y).value > 1:
    #                 ws.cell(row=x, column=y).fill = pink_fill
    #             elif ws.cell(row=x, column=y).value > 0:
    #                 ws.cell(row=x, column=y).fill = light_pink_fill
    #             elif ws.cell(row=x, column=y).value > -1:
    #                 ws.cell(row=x, column=y).fill = light_blue_fill
    #             elif ws.cell(row=x, column=y).value > -2:
    #                 ws.cell(row=x, column=y).fill = blue_fill
    #             elif ws.cell(row=x, column=y).value > -3:
    #                 ws.cell(row=x, column=y).fill = blue_fill_2
    #             elif ws.cell(row=x, column=y).value > -4:
    #                 ws.cell(row=x, column=y).fill = dark_blue_fill
    #             else:
    #                 ws.cell(row=x, column=y).fill = navy_fill
    #
    #     # 전체 보더
    #     for x in range(1, ws.max_row + 1):
    #         for y in range(1, ws.max_column + 1):
    #             ws.cell(row=x, column=y).border = thin_border
    #
    #     # 오토필터
    #     ws.auto_filter.ref = ws.dimensions
    #     ws.freeze_panes = 'C2'
    #
    #     # 하이퍼링크
    #     hyper_dict = {"업종": "https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no=",
    #                   "테마": "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no="}
    #     for i in range(2, ws.max_row):
    #         ws[f"B{i}"].hyperlink = hyper_dict[f"{item2}"] + str(ws[f"A{i}"].value)
    #
    # wb.save("stock_data/데일리등락률_통합.xlsx")
    # print("엑셀 서식 완료")

def create_total_data_excel():

    import pandas as pd
    import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
    pd.options.display.max_columns = 50
    pd.options.display.max_rows = 500

    ################
    # 1. 파일합치기
    ################
    writer = pd.ExcelWriter(f"stock_data/통합데이터/주식정보_통합_{datetime.date.today()}.xlsx")
    # 재무데이터
    for period in ['연간', '분기']:
        df = pd.read_excel(f'stock_data/재무실적/로우데이터/{period}실적_통합_2021-04-01.xlsx')
        print(f"{period} 실적 저장")
        df.to_excel(writer, sheet_name=f'{period}실적', index=None)

    # 등락률
    for item in ['업종', '테마']:
        fluc_df = pd.read_excel("stock_data/데일리등락률_통합.xlsx", sheet_name=f'{item}')
        fluc_df.to_excel(writer, sheet_name=f'{item}등락률', index=None)
        print(f"{item} 등락률 저장")

    # 실시간밸류
    value_df = pd.read_excel("stock_data/실시간벨류에이션/realtime_stock_value.xlsx")
    value_df.to_excel(writer, sheet_name="실시간밸류", index=None)
    print("실시간밸류 저장")
    writer.save()

    ###############
    # 2. 엑셀다듬기
    ###############
    # 셀서식
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                         bottom=Side(style="thin"))

    # 글꼴 서식
    font_normal = Font(name="맑은 고딕", size=11)
    font_bold = Font(name="맑은 고딕", size=11, bold=True)
    font_red = Font(name="맑은 고딕", size=11, color="FE2E2E")
    font_blue = Font(name="맑은 고딕", size=11, color="1E88E5")

    # color_fill
    orange_fill = PatternFill(patternType="solid", fgColor=Color("FEF5E7"))
    green_fill = PatternFill(patternType="solid", fgColor=Color("BBE8A2"))
    dark_orange_fill = PatternFill(patternType="solid", fgColor=Color("FC9C30"))

    light_pink_fill = PatternFill(patternType="solid", fgColor=Color("ffe6e6"))
    pink_fill = PatternFill(patternType="solid", fgColor=Color("ffcccc"))
    pink_fill_2 = PatternFill(patternType="solid", fgColor=Color("ffb3b3"))
    pink_fill_3 = PatternFill(patternType="solid", fgColor=Color("ff8080"))
    dark_pink_fill = PatternFill(patternType="solid", fgColor=Color("ff6666"))
    red_fill = PatternFill(patternType="solid", fgColor=Color("ff0000"))

    light_blue_fill = PatternFill(patternType="solid", fgColor=Color("e6f5ff"))
    blue_fill = PatternFill(patternType="solid", fgColor=Color("ccebff"))
    blue_fill_2 = PatternFill(patternType="solid", fgColor=Color("99d6ff"))
    blue_fill_3 = PatternFill(patternType="solid", fgColor=Color("66c2ff"))
    dark_blue_fill = PatternFill(patternType="solid", fgColor=Color("4db8ff"))
    navy_fill = PatternFill(patternType="solid", fgColor=Color("008ae6"))

    # 통합 데이터 엑셀 불러오기
    wb = load_workbook(f"stock_data/통합데이터/주식정보_통합_{datetime.date.today()}.xlsx")

    # 재무
    col_dict = {'연간': green_fill, '분기': dark_orange_fill}
    for item2 in col_dict:

        ws = wb[f'{item2}실적']

        # 헤더 서식
        for y in range(1, ws.max_column + 1):
            ws.cell(row=1, column=y).font = font_bold
            ws.cell(row=1, column=y).alignment = align_center
            ws.cell(row=1, column=y).fill = col_dict[item2]

        # 오토필터
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'D2'
        print(f"{item2} 엑셀 서식 완료")

    # 등락률
    for item3 in ["업종", "테마"]:
        ws = wb[f'{item3}등락률']

        # 컬럼너비
        ws.column_dimensions["B"].width = 30

        # 헤더 서식
        for y in range(1, ws.max_column + 1):
            ws.cell(row=1, column=y).font = font_bold
            ws.cell(row=1, column=y).alignment = align_center
            ws.cell(row=1, column=y).fill = orange_fill

        # 색칠
        for x in range(2, ws.max_row + 1):
            for y in range(3, ws.max_column + 1):
                if ws.cell(row=x, column=y).value == None:
                    continue
                if ws.cell(row=x, column=y).value > 5:
                    ws.cell(row=x, column=y).fill = red_fill
                elif ws.cell(row=x, column=y).value > 4:
                    ws.cell(row=x, column=y).fill = dark_pink_fill
                elif ws.cell(row=x, column=y).value > 3:
                    ws.cell(row=x, column=y).fill = pink_fill_3
                elif ws.cell(row=x, column=y).value > 2:
                    ws.cell(row=x, column=y).fill = pink_fill_2
                elif ws.cell(row=x, column=y).value > 1:
                    ws.cell(row=x, column=y).fill = pink_fill
                elif ws.cell(row=x, column=y).value > 0:
                    ws.cell(row=x, column=y).fill = light_pink_fill
                elif ws.cell(row=x, column=y).value > -1:
                    ws.cell(row=x, column=y).fill = light_blue_fill
                elif ws.cell(row=x, column=y).value > -2:
                    ws.cell(row=x, column=y).fill = blue_fill
                elif ws.cell(row=x, column=y).value > -3:
                    ws.cell(row=x, column=y).fill = blue_fill_2
                elif ws.cell(row=x, column=y).value > -4:
                    ws.cell(row=x, column=y).fill = dark_blue_fill
                else:
                    ws.cell(row=x, column=y).fill = navy_fill

        # 전체 보더
        for x in range(1, ws.max_row + 1):
            for y in range(1, ws.max_column + 1):
                ws.cell(row=x, column=y).border = thin_border

        # 오토필터
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'C2'

        # 하이퍼링크
        hyper_dict = {"업종": "https://finance.naver.com/sise/sise_group_detail.nhn?type=upjong&no=",
                      "테마": "https://finance.naver.com/sise/sise_group_detail.nhn?type=theme&no="}
        for i in range(2, ws.max_row):
            ws[f"B{i}"].hyperlink = hyper_dict[f"{item3}"] + str(ws[f"A{i}"].value)

        print(f"{item3} 엑셀 서식 완료")

    # 실시간 밸류
    ws = wb['실시간밸류']

    # 컬럼너비 (종목명, 업종명, 상승종목 수, 하락종목 수)
    width_dict = {"C": 21, "E": 19, "O": 13, "P": 13, "Q": 13}
    for key, value in width_dict.items():
        ws.column_dimensions[f"{key}"].width = value

    # 헤더 서식
    for row in ws["A1:M1"]:
        for cell in row:
            cell.font = font_bold
            cell.alignment = align_center
            cell.fill = orange_fill
            cell.border = thin_border

    # 등락률 글자색
    count_inc = 0  # 상승 종목 갯수
    count_dec = 0  # 하락 종목 갯수
    for col in ws[f"F2:F{ws.max_row}"]:
        for cell in col:
            if cell.value > 0:
                cell.font = font_red
                count_inc += 1
            else:
                cell.font = font_blue
                count_dec += 1

    # 하이퍼링크
    for i in range(2, ws.max_row):
        ws[f"C{i}"].hyperlink = ws[f"A{i}"].value

    # 상승 하락 종목 셀 서식
    ws["O1"].value = "상승 종목 수"
    ws["P1"].value = "하락 종목 수"
    ws["Q1"].value = "총 종목 수"
    ws["O1"].fill = dark_pink_fill
    ws["P1"].fill = dark_blue_fill
    ws["Q1"].fill = orange_fill
    ws["O2"].value = count_inc
    ws["P2"].value = count_dec
    ws["Q2"].value = count_inc + count_dec

    for row in ws[f"O1:Q1"]:
        for cell in row:
            cell.font = font_bold

    for row in ws[f"O1:Q2"]:
        for cell in row:
            cell.border = thin_border
            cell.alignment = align_center

    print(count_inc, count_dec, count_inc + count_dec)

    # 전체 보더
    for row in ws[f"A2:M{ws.max_row}"]:
        for cell in row:
            cell.border = thin_border

    # 오토필터
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    # 컬럼 숨기기 (URL)
    ws.column_dimensions["A"].hidden = True
    print("밸류 엑셀 서식 완료")

    wb.save(f"stock_data/통합데이터/주식정보_통합_{datetime.date.today()}.xlsx")


'''
6.etf/etn
'''
# ETF 코드 받기
# wise_company 사이트에서 etf 코드 가져오기
def get_etf_code_from_wise():
    from scraping.web_scraping import create_soup

    soup = create_soup("http://comp.wisereport.co.kr/ETF/lookup.aspx")
    table = soup.select_one('table.table').tbody.select('tr')

    etf_code = []
    for tag in table:
        td = tag.select('td')
        if td[0].text == "ETF":
            etf_code.append(td[1].text.strip())
        # else:
        #     etf_list.append([data.text.strip() for data in td])
    print("ETF code done")
    return etf_code

# ETN 코드받기
def get_etn_code_from_wise():
    from scraping.web_scraping import create_soup

    soup = create_soup("http://comp.wisereport.co.kr/ETF/lookup.aspx")
    table = soup.select_one('table.table').tbody.select('tr')

    etn_code = []
    for tag in table:
        td = tag.select('td')
        if td[0].text == "ETN":
            etn_code.append(td[1].text.strip())
        # else:
        #     etf_list.append([data.text.strip() for data in td])
    print("ETN code done")
    return etn_code

# 네이버 금융에서 개별 etf 정보 가져오기
# 작동 오류 남
def get_etf_info_from_naver():
    import time
    import datetime
    import csv
    from scraping.web_scraping import create_soup
    from selenium import webdriver
    from bs4 import BeautifulSoup

    # # CSV
    csv_file = open(f'stock_data\\etf_{datetime.date.today()}.csv', 'w', encoding='utf-8-sig', newline='')
    csv_writer = csv.writer(csv_file)
    csv_writer.writerow(['ETF이름', '링크', 'ETF코드', '운용사', '수수료', '시가총액(억 원)', '수익률', '', '', '', '구성종목 TOP 10'])
    csv_writer.writerow(
        ['', '', '', '', '', '', '1개월', '3개월', '6개월', '1년', '1위', '', '2위', '', '3위', '', '4위', '', '5위', '', '6위', '',
         '7위', '', '8위', '', '9위', '', '10위', ''])

    # ETF 코드
    etf_code = get_etf_code_from_wise()

    # 헤드리스 셀레니움
    options = webdriver.ChromeOptions()
    options.headless = True
    options.add_argument("window-size=1920x1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36")
    driver = webdriver.Chrome(options=options)

    print(f"Get ETF info :")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))

    for code in etf_code:
        print(f"Start : {code}")
        url = f'https://finance.naver.com/item/coinfo.nhn?code={code}'
        driver.get(url)
        driver.implicitly_wait(3)
        soup = create_soup(url)

        # 이름, 링크, 운용사, 수수료
        etf_name = soup.select_one('#middle > div.h_company > div.wrap_company > h2 > a').text
        etf_link = f'https://finance.naver.com/item/coinfo.nhn?code={code}'
        company = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[1].text
        commission = soup.select_one('#tab_con1').select_one('div:nth-child(4)').select('td')[0].em.text.replace('%',
                                                                                                                 '')
        each_etf_info = [etf_name, etf_link, str(code), company, commission]  # 1차 리스트 저장

        # iframe으로 이동
        driver.switch_to.frame('coinfo_cp')
        time.sleep(1)
        soup_iframe = BeautifulSoup(driver.page_source, "lxml")

        # 시가총액
        capital = soup_iframe.select_one('#status_grid_body').select('tr')[5].td.text.replace(',', '')[:-2]
        each_etf_info.append(capital)

        # 수익률
        earnings = soup_iframe.select_one('#status_grid_body').select('tr')[-1].td.select('span')
        for line in earnings:
            earning = line.text.replace('%', '').replace('+', '')
            each_etf_info.append(earning)

        # 구성종목
        top_10 = soup_iframe.select_one('#CU_grid_body').select('tr')[:10]
        for line in top_10:
            sh_name = line.select('td')[0].text
            percent = line.select('td')[2].text
            if percent == '-':
                percent = ''
            else:
                percent = percent
            each_etf_info.extend([sh_name, percent])
            # 리스트 진짜 이방법밖에 없는거 실화..? 더 찾아보기
            # each_etf_info.append(sh_name)  # 3차 리스트 저장
            # each_etf_info.append(percent)
        # csv에 입력
        csv_writer.writerow(each_etf_info)

    driver.quit()
    print(f"ETF info completed")
    print(time.strftime("%Y-%m-%d %H:%M:%S"))